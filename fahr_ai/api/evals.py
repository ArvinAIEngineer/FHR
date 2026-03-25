import requests
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime
from difflib import SequenceMatcher
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse
import tempfile
import shutil
import time
os.environ['NO_PROXY'] = '10.254.115.17,10.254.140.69'
# Create FastAPI app
app = FastAPI()

# API config
BASE_URL = "http://10.254.115.17:8090"
CHAT_URL = f"{BASE_URL}/api/Conversations/Conversation"

# Excel columns to be present in the output
OUTPUT_COLUMNS = [
    "Test Case",
    "Avatar",
    "Language",
    "Conv ID",
    "PERSON ID",
    "Expected Output",
    "AI Output",
    "Similarity Score",
    "Passed"
]


def send_conversation(personId, language, avatarId, content):
    """Send a conversation message to the API."""
    print(
        f"Making API call for Person ID: {personId}, Avatar: {avatarId}, Language: {language}"
    )
    print(content)
    headers = {"accept": "text/plain", "Content-Type": "application/json-patch+json"}

    payload = {
        "conversationId": 77377,
        "personId": personId,
        "avatarId": avatarId,
        "sessionStart": False,
        "language": language,
        "conversationMessage": content,
        "channel": "PRIVATE",
        "inputType": "TEXT",
        "outputType": "TEXT",
        "attachments": [
            {
                "fileName": "string",
                "filePath": "string",
                "contentType": "string",
                "fileContent": "string",
                "fileSize": 0,
            }
        ],
    }

    max_retries = 3
    timeout = 120

    for attempt in range(max_retries):
        try:
            print(f"Attempt {attempt + 1}/{max_retries} - Waiting for API response...")
            response = requests.post(
                CHAT_URL, json=payload, headers=headers, timeout=timeout
            )
            response.raise_for_status()
            result = response.json()
            print(result)
            print(
                f"API call successful - Conversation ID: {result.get('data', {}).get('conversationId', 'N/A')}"
            )
            return result
        except requests.exceptions.Timeout:
            print(f"Timeout on attempt {attempt + 1}")
            if attempt < max_retries - 1:
                print("Retrying in 2 seconds...")
                time.sleep(2)
                continue
            else:
                print("All retry attempts failed due to timeout")
                raise Exception(f"API timeout after {max_retries} attempts")
        except requests.exceptions.RequestException as e:
            print(f"Request error on attempt {attempt + 1}: {str(e)}")
            if attempt < max_retries - 1:
                print("Retrying in 2 seconds...")
                time.sleep(2)
                continue
            else:
                print("All retry attempts failed")
                raise Exception(f"API error: {str(e)}")



def similarity(a, b):
    """Calculate similarity between two strings."""
    if not a or not b:
        return 0.0
    score = round(SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio(), 2)
    print(f"Similarity score: {score}")
    return score

def similarity_embedding(expected, ai_output):
    import json
    import logging
    import subprocess
    from typing import List

    import numpy as np
    import pandas as pd

    # --- Configuration ---
    OLLAMA_BASE_URL = "http://10.254.140.69:11434"
    EMBEDDING_MODEL = "paraphrase-multilingual:latest"

    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)

    def get_embedding(text: str) -> List[float]:
        """Get embedding from Ollama"""
        try:
            result = subprocess.run(
                [
                    "curl",
                    "-s",
                    "-X",
                    "POST",
                    f"{OLLAMA_BASE_URL}/api/embeddings",
                    "-H",
                    "Content-Type: application/json",
                    "-d",
                    json.dumps({"model": EMBEDDING_MODEL, "prompt": text}),
                ],
                capture_output=True,
                text=True,
                timeout=30,
            )
            if result.returncode != 0:
                raise Exception("Curl failed")
            response = json.loads(result.stdout)
            embedding = response["embedding"]
            if len(embedding) != 768:
                raise Exception(f"Wrong embedding dimension: {len(embedding)}")
            return embedding
        except Exception as e:
            logger.error(f"Embedding failed: {e}")
            raise

    def cosine_similarity(vec1, vec2):
        vec1 = np.array(vec1)
        vec2 = np.array(vec2)
        if np.linalg.norm(vec1) == 0 or np.linalg.norm(vec2) == 0:
            return 0.0
        return float(np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2)))

    similarity_scores = []
    passed = []
    if not ai_output.strip():
        similarity_scores.append(np.nan)
        passed = "Skipped"
    try:
        emb_expected = get_embedding(expected)
        emb_ai = get_embedding(ai_output)
        sim = cosine_similarity(emb_expected, emb_ai)
        similarity_scores.append(sim)
        passed.append("Yes" if sim >= 0.70 else "No")
    except Exception as e:
        similarity_scores.append(np.nan)
        passed.append(f"Error: {e}")
    return passed, similarity_scores


def similarity_llm_grad(expected, ai_output):
    """Get LLM response using curl to Ollama"""
    import json
    import logging
    import subprocess
    from typing import List
    import re
    import numpy as np
    import pandas as pd
    def extract_final_answer(text):
        # Remove <think>...</think> block
        text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL)
        # Extract 'Yes' or 'No'
        match = re.search(r"\b(Yes|No)\b", text.strip())
        return match.group(1) if match else None
    # --- Configuration ---
    OLLAMA_BASE_URL = "http://10.254.140.69:11434"
    LLM_MODEL = "qwen3:32b"
    prompt = f"""Compare the expected output and the AI output. If the AI output conveys the same meaning or answer as the expected output, even if the wording is different, respond with only:
    'Yes'
    Otherwise, respond with only:
    'No'

    Expected: {expected}
    AI Output: {ai_output}

    Rule:
    - Don't add anything other than 'Yes' or 'No'."""
    try:
        llm_curl = [
            "curl",
            "-s",
            "-X",
            "POST",
            f"{OLLAMA_BASE_URL}/api/generate",
            "-H",
            "Content-Type: application/json",
            "-d",
            json.dumps({"model": LLM_MODEL, "prompt": prompt, "stream": False}),
        ]

        result = subprocess.run(llm_curl, capture_output=True, text=True, timeout=90)
        if result.returncode != 0:
            raise RuntimeError(f"Curl failed: {result.stderr}")

        llm_response = json.loads(result.stdout)
        response_text = llm_response["response"]
        print(f"Generated LLM response with length: {len(response_text)}")
        final_text=extract_final_answer(response_text)
        if final_text:
            return final_text
        return response_text

    except Exception as e:
        print(f"LLM generation failed: {e}")
        response_text = ""
        return response_text
@app.post("/evaluate-excel")
async def evaluate_excel(
    language: str = Form(...),
    file: UploadFile = File(...)
):
    """Evaluate uploaded Excel file and return updated file."""
    print("Starting Excel evaluation process...")
    print(f"Processing file: {file.filename}")
    print(f"Language: {language}")
    
    # Step 1: Save uploaded file to temp location
    print("Step 1: Saving uploaded file to temporary location...")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name
    print(f"File saved to: {tmp_path}")

    print("Creating Output File")
    output_path="/home/adminai/doaa_workspace/FAHR/fahr_ai/tests/ar_en_testcases_output_40Q_17-july.xlsx"
    wb_output = Workbook()
    ws_output = wb_output.active
    output_headers = [
        "Test Case",
        "Question",
        "Avatar",
        "Language",
        "Expected Output",
        "AI Output",
        "Similarity Score",
        "Passed",
        "Passed_llm_graded",
        "Estimated time",
        "chunks"
    ]
    ws_output.append(output_headers)
    # Step 2: Read Excel file
    print("Step 2: Reading Excel file...")
    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active
    total_rows = ws.max_row - 1  # Subtract header row
    print(f"Excel loaded - {total_rows} data rows found")
    headers = [cell.value for cell in ws[1]]
    print(headers)
    # if "Passed_llm_graded" not in headers:
    #     print("adding column")
    #     ws.cell(row=1, column=len(headers) + 1, value="Passed_llm_graded")
    #     for row in range(2, ws.max_row + 1):
    #         ws.cell(
    #             row=row, column=len(headers) + 1, value=None
    #         )  # Optional: fill with default
    #     wb.save(tmp_path)
    # Step 3: Process each row
    print("Step 3: Processing rows...")
    processed_count = 0
    error_count = 0
    
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        test_case = row[0] if len(row) > 0 else f"Case-{i}"
        content = row[1] if len(row) > 1 else ""
        person_id = row[5] if len(row) > 5 else ""  # Get person ID from Excel
        expected_output = row[6] if len(row) > 6 else ""
        avatarId = row[2] if len(row) > 6 else ""
        # language = row[3] if len(row) > 6 else ""

        print(f"\nRow {i-1}/{total_rows}: Test Case '{test_case}'")

        # Skip rows with empty questions
        if not content or str(content).strip() == "":
            # print("Skipping empty question row")
            continue

        # Step 4: Call API
        print("Step 4: Calling AI API...")
        try:
            start=time.time()
            response = send_conversation(person_id, language, avatarId, content)
            data = response.get("data", {})
            ai_text = data.get("textData", "")
            chunks =  str(data.get("referenceData", ""))
            if isinstance(chunks, list):
                chunks = ""
            # conv_id = data.get("conversationId", "")

            # Step 5: Calculate similarity and update
            print("Step 5: Calculating similarity and updating results...")
            # score = similarity(expected_output, ai_text)
            # passed = "Yes" if score >= 0.7 else "No"

            passed, score = similarity_embedding(expected_output, ai_text)
            passed_llm_grad = similarity_llm_grad(expected_output, ai_text)
            print("[Embedding]",passed[0],score[0])
            print("[LLM]",passed_llm_grad)
            time_taken = time.time() - start
            # Update worksheet
            output_row = [
                        test_case,
                        content,
                        avatarId,
                        language,
                        expected_output,
                        ai_text,
                        score[0],
                        passed[0],
                        passed_llm_grad,
                        time_taken,
                        chunks
                    ]
            ws_output.append(output_row)
            processed_count += 1

            # Save after each row so output file updates live
            wb_output.save(output_path)

        except Exception as e:
            print(f"Error in row {i-1}: {str(e)}")
            error_row = [
                test_case,
                content,
                0,
                language,
                expected_output,
                f"Error: {str(e)}",
                0,
                "No",
                "No",
                0,
            ]
            ws_output.append(error_row)
            error_count += 1
            wb_output.save(output_path)
    # Step 6: Save and return
    print(f"\nStep 6: Saving results...")
    print(f"Summary: {processed_count} rows processed successfully, {error_count} errors")
    wb_output.save(output_path)
    print(f"Results saved to: {output_path}")
    print("Evaluation process completed!")

    # Return file response
    return FileResponse(
        tmp_path,
        filename="evaluated.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# if __name__=="__main__":
#     a,b=similarity_embedding("I am good","I am good")
#     # a=similarity_llm_grad("I am good","I not am good")

#     print(a[0],b[0])
#/tmp/tmpiyoa158i.xlsx