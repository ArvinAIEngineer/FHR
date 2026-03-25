import sys
sys.path.append("./")
import openpyxl
from openpyxl import Workbook
import os
import json
import asyncio
from datetime import datetime
from difflib import SequenceMatcher
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse
import tempfile
import shutil
import time
import math
from orchestrator.orchestrator import Orchestrator
import zipfile

# Create FastAPI app
app = FastAPI()

# Excel columns for output
OUTPUT_COLUMNS = [
    "Test Case", "Content", "Avatar", "Language", "Conv ID", 
    "Person ID", "Expected Output", "AI Output", "Similarity Score", "Passed", "Processing Time (s)", "chunks"
]

# API config for conversation history
API_CONFIG = {
    "base_url": "http://10.254.115.17:8090",
    "headers": {
        "Content-Type": "application/json"
    },
    "timeout": 30
}

# Default batch size
DEFAULT_BATCH_SIZE = 10
# Initialize orchestrator (uncomment and adjust import)
orchestrator = Orchestrator()

async def call_orchestrator(person_id, language, avatar_id, content, conversation_id, user_info):
    """Call the orchestrator function directly."""
    
    # Prepare input data structure
    input_data = {
        "userId": person_id,
        "conversationId": conversation_id,
        "avatarId": avatar_id,
        "sessionStart": False,
        "conversationTitle": "Evaluation Test",
        "inputType": "TEXT",
        "outputType": "TEXT",
        "channel": "PRIVATE",
        "personId": person_id,
        "conversationMessage": content,
        "role": "EMPLOYEE",
        "attachments": [],
        "personalInfo": user_info
    }
    
    start_time = time.time()
    
    try:
        print(f"Calling orchestrator for Person {person_id}")
        
        # Call orchestrator directly (uncomment when orchestrator is imported)
        result = await orchestrator.run(input_data, API_CONFIG)
        
        processing_time = round(time.time() - start_time, 2)
        
        # Extract text response from orchestrator output
        ai_output = ""
        if result and len(result) > 0:
            for output in result:
                if output.get("type") == "text":
                    ai_output = output.get("data", "")
                    chunks =  str(output.get("referenceData", ""))
                    if isinstance(chunks, list):
                        chunks = ""
                    break
        
        print(f"Success - Time: {processing_time}s")
        return {
            "data": {
                "textData": ai_output,
                "chunks": chunks,
                "conversationId": conversation_id
            },
            "processing_time": processing_time
        }
        
    except Exception as e:
        processing_time = round(time.time() - start_time, 2)
        print(f"Error: {str(e)}, Time: {processing_time}s")
        raise Exception(f"Orchestrator failed: {str(e)}, Time: {processing_time}s")

def calculate_similarity(expected, actual):
    """Calculate similarity score between expected and actual output."""
    if not expected or not actual:
        return 0.0
    return round(SequenceMatcher(None, str(expected).lower(), str(actual).lower()).ratio(), 2)

async def process_batch(batch_data, avatarId, language, batch_num, conversation_id, user_info):
    """Process a single batch of test cases."""
    print(f"\n=== Processing Batch {batch_num} ({len(batch_data)} items) ===")
    
    results = []
    for i, row_data in enumerate(batch_data, 1):
        test_case = row_data.get('test_case', f"Case-{batch_num}-{i}")
        content = row_data.get('content', '')
        person_id = row_data.get('person_id', '')
        expected_output = row_data.get('expected_output', '')
        
        print(f"Batch {batch_num} - Item {i}/{len(batch_data)}: {test_case}")
        
        if not content.strip():
            print("Skipping empty content")
            continue
            
        try:
            # Call orchestrator directly
            response = await call_orchestrator(
                person_id, language, avatarId, content, conversation_id, user_info
            )
            data = response.get("data", {})
            ai_output = data.get("textData", "")
            conv_id = data.get("conversationId", "")
            chunks = data.get("chunks")
            processing_time = response.get("processing_time", 0.0)
            # Calculate similarity and pass/fail
            similarity_score = calculate_similarity(expected_output, ai_output)
            passed = "Yes" if similarity_score >= 0.7 else "No"
            
            results.append({
                'test_case': test_case,
                'content': content,
                'avatar': avatarId,
                'language': language,
                'conv_id': conv_id,
                'person_id': person_id,
                'expected_output': expected_output,
                'ai_output': ai_output,
                'similarity_score': similarity_score,
                'passed': passed,
                'processing_time': processing_time,
                'chunks': chunks
            })
            
            print(f"Completed - Score: {similarity_score}, Passed: {passed}, Time: {processing_time}s")
            
        except Exception as e:
            print(f"Error processing item {i}: {str(e)}")
            # Extract processing time from error message if available
            error_msg = str(e)
            processing_time = 0.0
            if "Time:" in error_msg:
                try:
                    processing_time = float(error_msg.split("Time: ")[1].split("s")[0])
                except:
                    processing_time = 0.0
            
            results.append({
                'test_case': test_case,
                'content': content,
                'avatar': avatarId,
                'language': language,
                'conv_id': '',
                'person_id': person_id,
                'expected_output': expected_output,
                'ai_output': f"Error: {str(e)}",
                'similarity_score': 0.0,
                'passed': 'No',
                'processing_time': processing_time,
                'chunks': chunks
            })
    
    print(f"Batch {batch_num} completed: {len(results)} items processed")
    return results

def create_batch_excel_file(batch_results, batch_num, output_dir, timestamp):
    """Create a separate Excel file for each batch."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch_{batch_num}"
    
    # Add headers
    for col, header in enumerate(OUTPUT_COLUMNS, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row, result in enumerate(batch_results, 2):
        ws.cell(row=row, column=1, value=result['test_case'])
        ws.cell(row=row, column=2, value=result['content'])
        ws.cell(row=row, column=3, value=result['avatar'])
        ws.cell(row=row, column=4, value=result['language'])
        ws.cell(row=row, column=5, value=result['conv_id'])
        ws.cell(row=row, column=6, value=result['person_id'])
        ws.cell(row=row, column=7, value=result['expected_output'])
        ws.cell(row=row, column=8, value=result['ai_output'])
        ws.cell(row=row, column=9, value=result['similarity_score'])
        ws.cell(row=row, column=10, value=result['passed'])
        ws.cell(row=row, column=11, value=result['processing_time'])
        ws.cell(row=row, column=12, value=result['chunks'])

    # Add summary statistics at the bottom
    summary_row = len(batch_results) + 3
    ws.cell(row=summary_row, column=1, value="BATCH SUMMARY")
    ws.cell(row=summary_row + 1, column=1, value="Total Items:")
    ws.cell(row=summary_row + 1, column=2, value=len(batch_results))
    
    passed_count = sum(1 for result in batch_results if result['passed'] == 'Yes')
    ws.cell(row=summary_row + 2, column=1, value="Passed:")
    ws.cell(row=summary_row + 2, column=2, value=passed_count)
    
    pass_rate = (passed_count / len(batch_results) * 100) if batch_results else 0
    ws.cell(row=summary_row + 3, column=1, value="Pass Rate:")
    ws.cell(row=summary_row + 3, column=2, value=f"{pass_rate:.1f}%")
    
    avg_score = sum(result['similarity_score'] for result in batch_results) / len(batch_results) if batch_results else 0
    ws.cell(row=summary_row + 4, column=1, value="Avg Similarity:")
    ws.cell(row=summary_row + 4, column=2, value=f"{avg_score:.2f}")
    
    total_time = sum(result['processing_time'] for result in batch_results)
    ws.cell(row=summary_row + 5, column=1, value="Total Time:")
    ws.cell(row=summary_row + 5, column=2, value=f"{total_time:.2f}s")
    
    # Save the file
    filename = f"batch_{batch_num:03d}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    print(f"Created batch file: {filename} with {len(batch_results)} rows")
    return filepath

def create_master_summary_file(all_results, output_dir, timestamp, avatar_id, language):
    """Create a master summary file with all results and overall statistics."""
    wb = Workbook()
    
    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Overall_Summary"
    
    # Overall statistics
    total_processed = sum(len(batch) for batch in all_results)
    total_passed = sum(1 for batch in all_results for result in batch if result['passed'] == 'Yes')
    avg_score = sum(result['similarity_score'] for batch in all_results for result in batch) / total_processed if total_processed > 0 else 0
    total_time = sum(result['processing_time'] for batch in all_results for result in batch)
    
    summary_ws.cell(row=1, column=1, value="EVALUATION SUMMARY")
    summary_ws.cell(row=2, column=1, value=f"Timestamp: {timestamp}")
    summary_ws.cell(row=3, column=1, value=f"Avatar ID: {avatar_id}")
    summary_ws.cell(row=4, column=1, value=f"Language: {language}")
    summary_ws.cell(row=5, column=1, value=f"Total Batches: {len(all_results)}")
    summary_ws.cell(row=6, column=1, value=f"Total Items: {total_processed}")
    summary_ws.cell(row=7, column=1, value=f"Total Passed: {total_passed}")
    summary_ws.cell(row=8, column=1, value=f"Overall Pass Rate: {total_passed/total_processed*100:.1f}%")
    summary_ws.cell(row=9, column=1, value=f"Average Similarity: {avg_score:.2f}")
    summary_ws.cell(row=10, column=1, value=f"Total Processing Time: {total_time:.2f}s")
    
    # Batch-by-batch summary
    summary_ws.cell(row=12, column=1, value="BATCH DETAILS")
    summary_ws.cell(row=13, column=1, value="Batch")
    summary_ws.cell(row=13, column=2, value="Items")
    summary_ws.cell(row=13, column=3, value="Passed")
    summary_ws.cell(row=13, column=4, value="Pass Rate")
    summary_ws.cell(row=13, column=5, value="Avg Similarity")
    summary_ws.cell(row=13, column=6, value="Time (s)")
    
    for i, batch_results in enumerate(all_results, 1):
        row = 13 + i
        batch_passed = sum(1 for result in batch_results if result['passed'] == 'Yes')
        batch_pass_rate = (batch_passed / len(batch_results) * 100) if batch_results else 0
        batch_avg_score = sum(result['similarity_score'] for result in batch_results) / len(batch_results) if batch_results else 0
        batch_time = sum(result['processing_time'] for result in batch_results)
        
        summary_ws.cell(row=row, column=1, value=f"Batch {i}")
        summary_ws.cell(row=row, column=2, value=len(batch_results))
        summary_ws.cell(row=row, column=3, value=batch_passed)
        summary_ws.cell(row=row, column=4, value=f"{batch_pass_rate:.1f}%")
        summary_ws.cell(row=row, column=5, value=f"{batch_avg_score:.2f}")
        summary_ws.cell(row=row, column=6, value=f"{batch_time:.2f}")
    
    # All results sheet
    all_ws = wb.create_sheet(title="All_Results")
    
    # Add headers
    for col, header in enumerate(OUTPUT_COLUMNS, 1):
        all_ws.cell(row=1, column=col, value=header)
    
    # Add all results
    row = 2
    for batch_results in all_results:
        for result in batch_results:
            all_ws.cell(row=row, column=1, value=result['test_case'])
            all_ws.cell(row=row, column=2, value=result['content'])
            all_ws.cell(row=row, column=3, value=result['avatar'])
            all_ws.cell(row=row, column=4, value=result['language'])
            all_ws.cell(row=row, column=5, value=result['conv_id'])
            all_ws.cell(row=row, column=6, value=result['person_id'])
            all_ws.cell(row=row, column=7, value=result['expected_output'])
            all_ws.cell(row=row, column=8, value=result['ai_output'])
            all_ws.cell(row=row, column=9, value=result['similarity_score'])
            all_ws.cell(row=row, column=10, value=result['passed'])
            all_ws.cell(row=row, column=11, value=result['processing_time'])
            all_ws.cell(row=row, column=12, value=result['chunks'])
            row += 1
    
    # Save the master file
    filename = f"master_summary_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    print(f"Created master summary file: {filename}")
    return filepath

def create_zip_archive(output_dir, timestamp):
    """Create a zip archive containing all batch files and summary."""
    zip_filename = f"evaluation_results_{timestamp}.zip"
    zip_filepath = os.path.join(output_dir, zip_filename)
    
    with zipfile.ZipFile(zip_filepath, 'w') as zipf:
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                if file.endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, output_dir)
                    zipf.write(file_path, arcname)
    
    print(f"Created zip archive: {zip_filename}")
    return zip_filepath

@app.post("/evaluate-batches")
async def evaluate_batches(
    avatarId: int = Form(...),
    language: str = Form(...),
    batch_size: int = Form(DEFAULT_BATCH_SIZE),
    conversation_id: int = Form(44344),
    user_info_file: str = Form("./tests/userInfo.json"),
    file: UploadFile = File(...)
):
    """Evaluate Excel file in batches using direct orchestrator calls."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    print(f"\n=== Starting Batch Evaluation ===")
    print(f"File: {file.filename}")
    print(f"Avatar ID: {avatarId}, Language: {language}, Batch Size: {batch_size}")
    print(f"Conversation ID: {conversation_id}")
    print(f"Timestamp: {timestamp}")
    
    # Create output directory
    output_dir = tempfile.mkdtemp()
    print(f"Output directory: {output_dir}")
    
    # Load user info
    user_info = {}
    try:
        if os.path.exists(user_info_file):
            with open(user_info_file, "r") as f:
                user_info = json.load(f)
            print(f"Loaded user info from {user_info_file}")
        else:
            print(f"User info file {user_info_file} not found, using empty user info")
    except Exception as e:
        print(f"Error loading user info: {str(e)}")
        user_info = {}
    
    # Save uploaded file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name
    
    # Read input Excel
    input_wb = openpyxl.load_workbook(tmp_path)
    input_ws = input_wb.active
    
    # Extract data from input file
    all_data = []
    for row in input_ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 1 and row[1]:  # Has content
            all_data.append({
                'test_case': row[0] if row[0] else f"Case-{len(all_data)+1}",
                'content': row[1],
                'person_id': row[5] if len(row) > 5 else '',
                'expected_output': row[6] if len(row) > 6 else ''
            })
    
    total_items = len(all_data)
    total_batches = math.ceil(total_items / batch_size)
    print(f"Total items: {total_items}, Batches: {total_batches}")
    
    # Process batches and create separate files
    all_results = []
    batch_files = []
    
    for batch_num in range(1, total_batches + 1):
        start_idx = (batch_num - 1) * batch_size
        end_idx = min(start_idx + batch_size, total_items)
        batch_data = all_data[start_idx:end_idx]
        
        # Process batch with orchestrator
        batch_results = await process_batch(
            batch_data, avatarId, language, batch_num, conversation_id, user_info
        )
        all_results.append(batch_results)
        
        # Create separate Excel file for this batch
        batch_file = create_batch_excel_file(batch_results, batch_num, output_dir, timestamp)
        batch_files.append(batch_file)
    
    # Create master summary file
    master_file = create_master_summary_file(all_results, output_dir, timestamp, avatarId, language)
    
    # Create zip archive with all files
    zip_file = create_zip_archive(output_dir, timestamp)
    
    # Calculate and display final summary
    total_processed = sum(len(batch) for batch in all_results)
    total_passed = sum(1 for batch in all_results for result in batch if result['passed'] == 'Yes')
    avg_score = sum(result['similarity_score'] for batch in all_results for result in batch) / total_processed if total_processed > 0 else 0
    
    print(f"\n=== Evaluation Complete ===")
    print(f"Total processed: {total_processed}")
    print(f"Total passed: {total_passed}")
    print(f"Pass rate: {total_passed/total_processed*100:.1f}%")
    print(f"Average similarity: {avg_score:.2f}")
    print(f"Generated {len(batch_files)} batch files + 1 master summary")
    print(f"All files packaged in: {os.path.basename(zip_file)}")
    
    # Clean up input file
    os.unlink(tmp_path)
    
    return FileResponse(
        zip_file,
        filename=f"evaluation_results_{timestamp}.zip",
        media_type="application/zip"
    )

# Standalone script version for direct execution
async def run_evaluation_standalone():
    """Run evaluation without FastAPI for direct testing."""
    
    # Configuration
    avatar_id = 0
    language = "en"
    batch_size = DEFAULT_BATCH_SIZE  # Now defaults to 20
    conversation_id = 663399
    user_info_file = "/home/adminai/doaa_workspace/FAHR/fahr_ai/tests/userInfo.json"
    input_excel = "/home/adminai/doaa_workspace/FAHR/fahr_ai/tests/AR_En_testcases_40Q_for_each.xlsx"  # Your input Excel file
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    print(f"\n=== Starting Standalone Batch Evaluation ===")
    print(f"Input file: {input_excel}")
    print(f"Avatar ID: {avatar_id}, Language: {language}, Batch Size: {batch_size}")
    print(f"Conversation ID: {conversation_id}")
    print(f"Timestamp: {timestamp}")
    
    # Create output directory
    output_dir = f"evaluation_results_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"Output directory: {output_dir}")
    
    # Load user info
    user_info = {}
    try:
        if os.path.exists(user_info_file):
            with open(user_info_file, "r") as f:
                user_info = json.load(f)
            print(f"Loaded user info from {user_info_file}")
        else:
            print(f"User info file {user_info_file} not found, using empty user info")
    except Exception as e:
        print(f"Error loading user info: {str(e)}")
        user_info = {}
    
    # Read input Excel
    input_wb = openpyxl.load_workbook(input_excel)
    input_ws = input_wb.active
    
    # Extract data from input file
    all_data = []
    for row in input_ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 1 and row[1]:  # Has content
            all_data.append({
                'test_case': row[0] if row[0] else f"Case-{len(all_data)+1}",
                'content': row[1],
                'person_id': row[5] if len(row) > 5 else '',
                'expected_output': row[6] if len(row) > 6 else ''
            })
    
    total_items = len(all_data)
    total_batches = math.ceil(total_items / batch_size)
    print(f"Total items: {total_items}, Batches: {total_batches}")
    
    # Process batches and create separate files
    all_results = []
    batch_files = []
    
    for batch_num in range(1, total_batches + 1):
        start_idx = (batch_num - 1) * batch_size
        end_idx = min(start_idx + batch_size, total_items)
        batch_data = all_data[start_idx:end_idx]
        
        # Process batch with orchestrator
        batch_results = await process_batch(
            batch_data, avatar_id, language, batch_num, conversation_id, user_info
        )
        all_results.append(batch_results)
        
        # Create separate Excel file for this batch
        batch_file = create_batch_excel_file(batch_results, batch_num, output_dir, timestamp)
        batch_files.append(batch_file)
    
    # Create master summary file
    master_file = create_master_summary_file(all_results, output_dir, timestamp, avatar_id, language)
    
    # Create zip archive with all files
    zip_file = create_zip_archive(output_dir, timestamp)
    
    # Calculate and display final summary
    total_processed = sum(len(batch) for batch in all_results)
    total_passed = sum(1 for batch in all_results for result in batch if result['passed'] == 'Yes')
    avg_score = sum(result['similarity_score'] for batch in all_results for result in batch) / total_processed if total_processed > 0 else 0
    
    print(f"\n=== Evaluation Complete ===")
    print(f"Total processed: {total_processed}")
    print(f"Total passed: {total_passed}")
    print(f"Pass rate: {total_passed/total_processed*100:.1f}%")
    print(f"Average similarity: {avg_score:.2f}")
    print(f"Generated {len(batch_files)} batch files + 1 master summary")
    print(f"All files available in: {output_dir}")
    print(f"Zip archive created: {os.path.basename(zip_file)}")
    
    # List all generated files
    print(f"\nGenerated files:")
    for file in os.listdir(output_dir):
        print(f"  - {file}")

if __name__ == "__main__":
    import uvicorn
    
    # Choose how to run:
    # 1. Run as FastAPI server
    # uvicorn.run(app, host="0.0.0.0", port=8000)
    
    # 2. Run standalone evaluation directly
    asyncio.run(run_evaluation_standalone())